#!/usr/bin/env python3
"""
Newsroom Tracker
----------------
Tracks press releases from multiple company newsrooms via Google News RSS
and reports any that are new since the last run.

Also tracks Blocks & Files (blocksandfiles.com), filtering articles to only
those that mention a tracked vendor by name in the title.

For each new article found, fetches the article content and uses the Claude API
to classify it as one of: New Product, New Feature, Partnership, Financial, Other.

Add or remove companies in the SOURCES list below.

State is stored in 'newsroom_seen_articles.json' in the output directory.
New articles are appended to 'newsroom_new_articles.csv' each run.
New articles are also written to 'newsroom_new_articles.xlsx' with one tab per vendor.
Full article list is written to 'newsroom_articles.csv' each run.

Usage:
    python3 manufacturers_news_tracker.py                     # no date filter
    python3 manufacturers_news_tracker.py --days 30           # only articles from last 30 days
    python3 manufacturers_news_tracker.py --since 2026-01-01  # only articles on/after this date

Requirements:
    pip3 install requests anthropic openpyxl
"""

import argparse
import csv
import json
import sys
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta, timezone
from pathlib import Path
from urllib.parse import urlencode

try:
    import requests
except ImportError:
    print("Missing dependency. Install with:")
    print("  pip3 install requests anthropic")
    sys.exit(1)

try:
    import anthropic
except ImportError:
    print("Missing dependency. Install with:")
    print("  pip3 install anthropic")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency. Install with:")
    print("  pip3 install openpyxl")
    sys.exit(1)

# ── Sources ────────────────────────────────────────────────────────────────────
# Add or remove entries here to track more/fewer newsrooms.
# Each entry needs:
#   "name"  - display label used in output
#   "query" - Google News search query (site: is the most reliable filter)

SOURCES = [
    {
        "name":  "NetApp",
        "query": "site:netapp.com/newsroom/press-releases",
    },
    {
        "name":  "Pure",
        "query": "site:purestorage.com/company/newsroom/press-releases",
    },
    {
        "name":  "Rubrik",
        "query": "site:rubrik.com/company/newsroom",
    },
    {
        "name":  "Cohesity",
        "query": "site:www.cohesity.com/newsroom/press",
    },
    {
        "name":  "Commvault",
        "query": "site:www.commvault.com/news",
    },
    {
        "name":  "Veeam",
        "query": "site:veeam.com/company/press-release",
    },
]

# ── Blocks & Files config ──────────────────────────────────────────────────────
# Direct RSS feed from blocksandfiles.com. Articles are filtered to only those
# whose title contains one of the vendor names from SOURCES (case-insensitive).
# Matched articles are labelled "Blocks & Files – <VendorName>".

BANDF_RSS_URL = "https://blocksandfiles.com/feed"

# ── Configuration ──────────────────────────────────────────────────────────────

OUTPUT_DIR   = Path("/Users/rick/Library/CloudStorage/OneDrive-Personal/Vendor Documentation/a_press_releases")
STATE_FILE   = OUTPUT_DIR / "newsroom_seen_articles.json"
CSV_FILE      = OUTPUT_DIR / "newsroom_new_articles.csv"
XLSX_FILE     = OUTPUT_DIR / "newsroom_new_articles.xlsx"
ALL_CSV_FILE  = OUTPUT_DIR / "newsroom_articles.csv"
ALL_XLSX_FILE = OUTPUT_DIR / "newsroom_articles.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)  # create if it doesn't exist

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/rss+xml, application/xml, text/xml, */*",
}

ARTICLE_FETCH_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

CATEGORIES = ["New Product", "New Feature", "Partnership", "Financial", "Other"]

# ── Date conversion ────────────────────────────────────────────────────────────

DATE_FORMATS = [
    "%a, %d %b %Y %H:%M:%S %Z",   # Wed, 16 Apr 2026 10:00:00 GMT
    "%a, %d %b %Y %H:%M:%S %z",   # Wed, 16 Apr 2026 10:00:00 +0000
    "%Y-%m-%dT%H:%M:%S%z",        # 2026-04-16T10:00:00+00:00
    "%Y-%m-%dT%H:%M:%SZ",         # 2026-04-16T10:00:00Z
    "%Y-%m-%d",                    # 2026-04-16
]

def parse_date(raw: str) -> str:
    """Convert a raw date string to YYYY-MM-DD, or return original if unparseable."""
    if not raw:
        return ""
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw.strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return raw


def parse_date_obj(raw: str) -> datetime | None:
    """Parse a raw date string into a timezone-aware datetime, or return None."""
    if not raw:
        return None
    for fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(raw.strip(), fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except ValueError:
            continue
    return None

# ── Argument parsing ───────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Check company newsrooms for new press releases."
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "--days",
        type=int,
        metavar="N",
        help="Only report articles published within the last N days (e.g. --days 30)",
    )
    group.add_argument(
        "--since",
        type=str,
        metavar="YYYY-MM-DD",
        help="Only report articles published on or after this date (e.g. --since 2026-01-01)",
    )
    return parser.parse_args()


def get_cutoff(args) -> datetime | None:
    """Return a timezone-aware cutoff datetime, or None if no filter was requested."""
    if args.days:
        return datetime.now(tz=timezone.utc) - timedelta(days=args.days)
    if args.since:
        try:
            dt = datetime.strptime(args.since, "%Y-%m-%d")
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            print(f"[ERROR] --since date '{args.since}' must be in YYYY-MM-DD format.")
            sys.exit(1)
    return None

# ── State helpers ──────────────────────────────────────────────────────────────

def load_seen() -> dict:
    if STATE_FILE.exists():
        try:
            with STATE_FILE.open() as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {}


def save_seen(seen: dict) -> None:
    with STATE_FILE.open("w") as f:
        json.dump(seen, f, indent=2)

# ── Article classification ─────────────────────────────────────────────────────

def fetch_article_text(url: str) -> str:
    """Fetch the text content of an article URL. Returns empty string on failure."""
    try:
        resp = requests.get(url, headers=ARTICLE_FETCH_HEADERS, timeout=15)
        resp.raise_for_status()
        # Strip HTML tags with a simple approach using xml/html parser
        import html
        import re
        text = resp.text
        # Remove script and style blocks
        text = re.sub(r'<(script|style)[^>]*>.*?</(script|style)>', '', text, flags=re.DOTALL | re.IGNORECASE)
        # Remove all remaining HTML tags
        text = re.sub(r'<[^>]+>', ' ', text)
        # Decode HTML entities and collapse whitespace
        text = html.unescape(text)
        text = re.sub(r'\s+', ' ', text).strip()
        # Return first 3000 chars — enough for classification
        return text[:3000]
    except Exception:
        return ""


def classify_article(title: str, url: str) -> str:
    """
    Use the Claude API to classify an article as one of the defined categories.
    Falls back to the article title alone if the URL can't be fetched.
    """
    body_text = fetch_article_text(url)

    if body_text:
        content = f"Title: {title}\n\nArticle text:\n{body_text}"
    else:
        content = f"Title: {title}"

    prompt = f"""You are classifying a technology press release or news article into exactly one of these categories:

- New Product: Announcement of a brand new product being launched
- New Feature: Announcement of a new feature, capability, or enhancement added to an existing product
- Partnership: Announcement of a partnership, integration, alliance, or collaboration with another company
- Financial: Financial results, earnings, funding rounds, acquisitions, or investor news
- Other: Anything that doesn't clearly fit the above (e.g. awards, executive appointments, events, surveys)

Article to classify:
{content}

Reply with ONLY the category name, exactly as written above. No explanation, no punctuation, nothing else."""

    try:
        client = anthropic.Anthropic()
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=20,
            messages=[{"role": "user", "content": prompt}]
        )
        result = message.content[0].text.strip()
        # Validate the response is one of our expected categories
        if result in CATEGORIES:
            return result
        # Try case-insensitive match as fallback
        for cat in CATEGORIES:
            if cat.lower() == result.lower():
                return cat
        return "Other"
    except Exception as e:
        print(f"      [WARNING] Classification failed: {e}")
        return "Unknown"

# ── CSV helpers ────────────────────────────────────────────────────────────────

def get_data_source(source: str) -> str:
    """Return 'Blocks and Files' or 'Google' based on the source field."""
    if source.startswith("Blocks & Files"):
        return "Blocks and Files"
    return "Google"


def get_clean_source(source: str) -> str:
    """Strip the 'Blocks & Files – ' prefix if present, returning just the vendor name."""
    if source.startswith("Blocks & Files – "):
        return source[len("Blocks & Files – "):]
    return source


def append_to_csv(new_articles: list[dict]) -> None:
    """Append new articles to the new-articles CSV, creating header if needed."""
    write_header = not CSV_FILE.exists()
    with CSV_FILE.open("a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["source", "data source", "date", "category", "title", "url"])
        if write_header:
            writer.writeheader()
        for a in new_articles:
            raw_source = a["source"]
            writer.writerow({
                "source":      get_clean_source(raw_source),
                "data source": get_data_source(raw_source),
                "date":        parse_date(a["date"]),
                "title":       a["title"],
                "url":         a["url"],
                "category":    a.get("category", ""),
            })


def write_new_articles_xlsx(new_articles: list[dict]) -> None:
    """
    Append new articles to a single-sheet Excel workbook.
    Headers are bold and light blue. Columns are auto-sized.
    URLs are written as clickable hyperlinks.
    If the file already exists it is loaded and rows are appended.
    """
    COLUMNS        = ["source", "data source", "date", "category", "title", "url"]
    HEADER_FONT    = Font(bold=True, color="000000")
    HEADER_FILL    = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    HYPERLINK_FONT = Font(color="0563C1", underline="single")
    SHEET_NAME     = "New Articles"

    # Load existing workbook or create a fresh one
    if XLSX_FILE.exists():
        wb = load_workbook(XLSX_FILE)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        next_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        next_row = 2
        # Write header row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.title())
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    # Sort newest first before appending
    articles = sorted(new_articles, key=lambda a: parse_date(a["date"]), reverse=True)

    # Write article rows
    for a in articles:
        raw_source = a["source"]
        row_data = [
            get_clean_source(raw_source),
            get_data_source(raw_source),
            parse_date(a["date"]),
            a.get("category", ""),
            a["title"],
            a["url"],
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=value)
            if COLUMNS[col_idx - 1] == "url" and value:
                cell.hyperlink = value
                cell.font = HYPERLINK_FONT
        next_row += 1

    # Auto-size all columns based on content
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

    wb.save(XLSX_FILE)


def write_all_articles_csv(seen: dict) -> None:
    """Write the full contents of the seen dict to newsroom_articles.csv."""
    rows = []
    for url, entry in seen.items():
        raw_source = entry.get("source", "")
        rows.append({
            "source":      get_clean_source(raw_source),
            "data source": get_data_source(raw_source),
            "date":        parse_date(entry.get("date", "")),
            "title":       entry.get("title", ""),
            "url":         url,
            "category":    entry.get("category", ""),
        })
    # Sort by source A-Z, then date newest to oldest within each source
    rows.sort(key=lambda r: r["date"], reverse=True)
    rows.sort(key=lambda r: r["source"].lower())
    with ALL_CSV_FILE.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["source", "data source", "date", "category", "title", "url"])
        writer.writeheader()
        writer.writerows(rows)


def write_all_articles_xlsx(seen: dict) -> None:
    """
    Write the full contents of the seen dict to newsroom_articles.xlsx.
    One tab per vendor, sorted newest first. Bold light blue headers,
    auto-sized columns, clickable URL hyperlinks.
    Overwrites the file completely each run (full snapshot).
    """
    from collections import defaultdict

    COLUMNS     = ["source", "data source", "date", "category", "title", "url"]
    HEADER_FONT = Font(bold=True, color="000000")
    HEADER_FILL = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    HYPERLINK_FONT = Font(color="0563C1", underline="single")

    # Build rows grouped by clean source name
    by_source = defaultdict(list)
    for url, entry in seen.items():
        raw_source = entry.get("source", "")
        by_source[get_clean_source(raw_source)].append({
            "source":      get_clean_source(raw_source),
            "data source": get_data_source(raw_source),
            "date":        parse_date(entry.get("date", "")),
            "category":    entry.get("category", ""),
            "title":       entry.get("title", ""),
            "url":         url,
        })

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for source_name in sorted(by_source.keys()):
        articles = sorted(by_source[source_name], key=lambda r: r["date"], reverse=True)
        ws = wb.create_sheet(title=source_name)

        # Header row
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.title())
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, a in enumerate(articles, start=2):
            row_data = [
                a["source"], a["data source"], a["date"],
                a["category"], a["title"], a["url"],
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if COLUMNS[col_idx - 1] == "url" and value:
                    cell.hyperlink = value
                    cell.font = HYPERLINK_FONT

        # Auto-size columns
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(col_name)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 80)

        # Add Excel AutoFilter dropdowns to all columns
        ws.auto_filter.ref = ws.dimensions

    wb.save(ALL_XLSX_FILE)

# ── RSS fetchers ───────────────────────────────────────────────────────────────

def rss_url(query: str) -> str:
    params = {"q": query, "hl": "en-US", "gl": "US", "ceid": "US:en"}
    return f"https://news.google.com/rss/search?{urlencode(params)}"


def fetch_feed(url: str, source_name: str) -> list[dict] | None:
    """Fetch any RSS feed URL and return raw items, or None on error."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  [ERROR] Could not fetch feed for {source_name}: {e}")
        return None
    try:
        root = ET.fromstring(resp.content)
    except ET.ParseError as e:
        print(f"  [ERROR] Could not parse feed for {source_name}: {e}")
        return None

    items = []
    for item in root.findall(".//item"):
        title = item.findtext("title", "").strip()
        url   = item.findtext("link",  "").strip()
        date  = item.findtext("pubDate", "").strip()
        if title and url:
            items.append({"title": title, "url": url, "date": date})
    return items


def fetch_google_news_articles(source: dict) -> list[dict]:
    """Fetch articles for a vendor via Google News RSS."""
    items = fetch_feed(rss_url(source["query"]), source["name"])
    if items is None:
        return []
    return [dict(item, source=source["name"]) for item in items]


def fetch_bandf_articles() -> list[dict]:
    """
    Fetch the Blocks & Files RSS feed and return only articles whose title
    contains at least one vendor name from SOURCES. Each matched article is
    labelled 'Blocks & Files – <VendorName>'.
    """
    print("── Blocks & Files ──")
    items = fetch_feed(BANDF_RSS_URL, "Blocks & Files")
    if items is None:
        print("   No articles found.\n")
        return []

    # Build lookup: lowercase vendor name -> display name
    vendor_names = {s["name"].lower(): s["name"] for s in SOURCES}

    matched = []
    for item in items:
        title_lower = item["title"].lower()
        for key, display_name in vendor_names.items():
            if key in title_lower:
                matched.append(dict(item, source=f"Blocks & Files – {display_name}"))
                break  # only label with the first matching vendor

    print(f"   {len(items)} article(s) in feed, {len(matched)} match a tracked vendor.")
    return matched

# ── Main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    args   = parse_args()
    cutoff = get_cutoff(args)

    cutoff_desc = ""
    if cutoff:
        cutoff_desc = f" (filtering articles on/after {cutoff.strftime('%Y-%m-%d')})"

    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Checking {len(SOURCES)} newsroom(s) + Blocks & Files{cutoff_desc} ...\n")

    seen         = load_seen()
    is_first_run = len(seen) == 0
    timestamp    = datetime.now().isoformat()

    total_new = 0
    all_new   = []

    def process_source(source_name: str, articles: list[dict]) -> None:
        nonlocal total_new

        if not articles:
            print("   No articles found.\n")
            return

        # Apply date cutoff filter if requested
        if cutoff:
            before   = len(articles)
            articles = [
                a for a in articles
                if (dt := parse_date_obj(a["date"])) is None or dt >= cutoff
            ]
            filtered = before - len(articles)
            if filtered:
                print(f"   {filtered} article(s) filtered out by date.")

        if is_first_run:
            for a in articles:
                seen[a["url"]] = {
                    "title":      a["title"],
                    "date":       a["date"],
                    "source":     a["source"],
                    "category":   "",
                    "first_seen": timestamp,
                }
            print(f"   Saved as baseline (first run).\n")
            return

        new_articles = [a for a in articles if a["url"] not in seen]

        if not new_articles:
            print("   No new articles since last check.\n")
        else:
            print(f"   {len(new_articles)} NEW article(s):\n")
            for i, a in enumerate(new_articles, 1):
                print(f"   {i}. {a['title']}")
                if a["date"]:
                    print(f"      Published: {parse_date(a['date'])}")
                print(f"      {a['url']}")
                print(f"      Classifying...", end=" ", flush=True)
                category = classify_article(a["title"], a["url"])
                a["category"] = category
                print(category)
                print()
                seen[a["url"]] = {
                    "title":      a["title"],
                    "date":       a["date"],
                    "source":     a["source"],
                    "category":   category,
                    "first_seen": timestamp,
                }
            all_new.extend(new_articles)
            total_new += len(new_articles)

    # Process each vendor newsroom
    for source in SOURCES:
        print(f"── {source['name']} ──")
        articles = fetch_google_news_articles(source)
        if articles:
            print(f"   {len(articles)} article(s) in feed.")
        process_source(source["name"], articles)

    # Process Blocks & Files
    bandf_articles = fetch_bandf_articles()
    if bandf_articles:
        print()
    process_source("Blocks & Files", bandf_articles)

    save_seen(seen)
    write_all_articles_csv(seen)
    write_all_articles_xlsx(seen)

    if is_first_run:
        print("First run complete — baselines saved.")
        print("Run the script again to start detecting new articles.")
    else:
        print(f"Done. {total_new} new article(s) found across all sources.")
        print(f"State saved to: {STATE_FILE}")
        print(f"Full article list updated: {ALL_CSV_FILE}")
        print(f"Full article list updated: {ALL_XLSX_FILE}")
        if all_new:
            append_to_csv(all_new)
            write_new_articles_xlsx(all_new)
            print(f"New articles appended to: {CSV_FILE}")
            print(f"New articles written to:  {XLSX_FILE}")


if __name__ == "__main__":
    main()
